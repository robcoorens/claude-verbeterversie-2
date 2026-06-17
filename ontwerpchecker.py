#!/usr/bin/env python3
"""
ontwerpchecker.py - voorbereiding + (optionele) review van grote secundaire
ontwerppakketten. Pure-Python motor op basis van PyMuPDF (geen externe
programma's nodig), zodat de tool tot een zelfstandige .exe te bundelen is.

Subcommando's:  triage | prepare | review | all
Zie README.md.  De webapp (webapp.py) roept do_prepare rechtstreeks aan.
"""
from __future__ import annotations
import argparse, glob, json, math, os, re, shutil, sys, textwrap
from pathlib import Path

VERSION = "2026-06-17"

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

DEFAULTS = {
    "max_file_mb": 28.0,
    "max_pages_per_pdf": 100,
    "max_files_per_chat": 20,
    "context_tokens": 200_000,
    "klassiek_drempel": 300,
    "raster_dpi": 100,
    "max_raster_pages": 60,
    "max_images": 20,             # Claude.ai: max 20 afbeeldingen per bericht
    "max_upload_files": 20,        # Claude.ai: max bestanden per gesprek
    "upload_chars_per_file": 500_000,
}

# Sleutelpagina-detectie: woordgrens-patronen (voorkomt valse 'ct'/'vt'-matches).
KEY_PATTERNS = [re.compile(p, re.I) for p in (
    r"\bgrondschema", r"\buitschakelmatrix", r"\buitschakel", r"\bfunctieschema",
    r"\bvergrendel", r"\bbeveiliging", r"\bstroom", r"\bspanning",
    r"\b(?:ct|vt)\d*\b", r"\boverdracht", r"\binschakeling", r"\bklemmenstrook",
    r"\bkastenbouw",
)]
# Sterkere, minder dubbelzinnige termen voor de tekst-terugval (anders te veel matches
# op brede woorden als 'stroom'/'spanning' die op vrijwel elke elektrische pagina staan).
KEY_PATTERNS_STRONG = [re.compile(p, re.I) for p in (
    r"\bgrondschema", r"\buitschakelmatrix", r"\bfunctieschema", r"\bvergrendel",
    r"\bklemmenstrook", r"\bkastenbouw",
)]


def _key_match(text, patterns):
    return any(p.search(text or "") for p in patterns)


# Gewogen termen: belangrijke tekeningtypes krijgen voorrang onder een krap budget.
TERM_WEIGHTS = [
    (re.compile(r"\bgrondschema", re.I), 5),
    (re.compile(r"\buitschakelmatrix", re.I), 5),
    (re.compile(r"\bfunctieschema", re.I), 4),
    (re.compile(r"\bvergrendel", re.I), 4),
    (re.compile(r"\bbeveiliging", re.I), 3),
    (re.compile(r"\b(?:ct|vt)\d*\b", re.I), 3),
    (re.compile(r"\boverdracht", re.I), 3),
    (re.compile(r"\binschakeling", re.I), 3),
    (re.compile(r"\buitschakel", re.I), 3),
    (re.compile(r"\bklemmenstrook", re.I), 2),
    (re.compile(r"\bkastenbouw", re.I), 2),
    (re.compile(r"\bstroom", re.I), 2),
    (re.compile(r"\bspanning", re.I), 2),
]


def _score_title(title):
    title = title or ""
    return sum(w for pat, w in TERM_WEIGHTS if pat.search(title))


def _page_bonus(page):
    """Grootformaat liggende pagina's zijn vrijwel altijd schema's -> voorrang."""
    w, h = page.rect.width, page.rect.height
    big = max(w, h) > 1600          # A2 en groter
    landscape = w > h
    return 2 if (big and landscape) else (1 if big else 0)


_OCR_OK = None


def _ocr_available():
    global _OCR_OK
    if _OCR_OK is None:
        _OCR_OK = shutil.which("tesseract") is not None
    return _OCR_OK


# --------------------------------------------------------------- infrastructuur
def app_dir():
    """Map met de meegeleverde bestanden (methodiek/, web/), ook in een .exe."""
    if getattr(sys, "frozen", False):
        return getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def die(msg, code=1):
    print(f"\nFOUT: {msg}", file=sys.stderr)
    sys.exit(code)


def need_fitz():
    if fitz is None:
        die("PyMuPDF ontbreekt. Installeer met:  pip install PyMuPDF")


def human_mb(path):
    return os.path.getsize(path) / (1024 * 1024)


def est_tokens(chars):
    return int(chars / 4)


def load_pages(txt_path):
    with open(txt_path, encoding="utf-8", errors="replace") as f:
        return f.read().split("\f")


# --------------------------------------------------------------- inventarisatie
def _page_size_label(page):
    w, h = page.rect.width, page.rect.height
    a = {"A4": (595, 842), "A3": (842, 1191), "A2": (1191, 1684),
         "A1": (1684, 2384), "A0": (2384, 3370)}
    lo, hi = sorted((w, h))
    for name, (aw, ah) in a.items():
        if abs(lo - aw) < 12 and abs(hi - ah) < 12:
            return f"{w:.0f} x {h:.0f} pts ({name})"
    return f"{w:.0f} x {h:.0f} pts"


def inventory(in_dir, cfg):
    need_fitz()
    pdfs = sorted(glob.glob(os.path.join(in_dir, "*.pdf")) +
                  glob.glob(os.path.join(in_dir, "*.PDF")))
    if not pdfs:
        die(f"Geen PDF's gevonden in {in_dir}")
    items, total = [], 0
    for pdf in pdfs:
        d = fitz.open(pdf)
        pages = d.page_count
        size = _page_size_label(d[0]) if pages else "?"
        sample = min(pages, 12)
        txt = sum(len(d[i].get_text("text")) for i in range(sample))
        has_text = (txt / max(1, sample)) > 40
        outline = len(d.get_toc())
        d.close()
        total += pages
        items.append({"path": pdf, "name": os.path.basename(pdf), "pages": pages,
                      "mb": round(human_mb(pdf), 1), "page_size": size,
                      "has_text": has_text, "outline": outline})
    mode = "klassiek" if total <= cfg["klassiek_drempel"] else "grootschalig"
    return {"pdfs": items, "total_pages": total, "mode": mode}


def decide_strategy(inv, cfg):
    per = []
    for it in inv["pdfs"]:
        r = []
        if it["pages"] > cfg["max_pages_per_pdf"]:
            r.append(f"{it['pages']} pagina's > {cfg['max_pages_per_pdf']} (PDF-limiet)")
        if it["mb"] > cfg["max_file_mb"]:
            r.append(f"{it['mb']} MB > {cfg['max_file_mb']} MB")
        if not it["has_text"]:
            r.append("geen tekstlaag (gescand -> OCR nodig)")
        per.append({**it, "direct_upload": not r, "redenen": r})
    n_direct = sum(1 for p in per if p["direct_upload"])
    est = est_tokens(inv["total_pages"] * 1800)
    if inv["mode"] == "klassiek" and n_direct == len(per):
        approach = "direct"
    elif est > cfg["context_tokens"] * 0.6 or inv["total_pages"] > 1500:
        approach = "veld_voor_veld"
    else:
        approach = "tekstpakket"
    return {"per_pdf": per, "n_direct": n_direct, "n_split": len(per) - n_direct,
            "approach": approach, "est_corpus_tokens": est}


def render_triage(inv, strat, cfg):
    L, w = [], None
    out = []
    def w(s): out.append(s)
    w("# Triagerapport - secundair ontwerppakket\n")
    w(f"- Bestanden: **{len(inv['pdfs'])}**")
    w(f"- Pagina's totaal: **{inv['total_pages']}**")
    w(f"- Modus: **{inv['mode'].upper()}**")
    w(f"- Geschat tekstcorpus: ~{strat['est_corpus_tokens']:,} tokens "
      f"(context ~{cfg['context_tokens']:,})\n")
    w("## Per bestand\n")
    w("| Bestand | Pag. | MB | Tekstlaag | Outline | Direct? | Reden |")
    w("|---|---:|---:|:--:|---:|:--:|---|")
    for p in strat["per_pdf"]:
        w(f"| {p['name']} | {p['pages']} | {p['mb']} | "
          f"{'ja' if p['has_text'] else 'NEE'} | {p['outline']} | "
          f"{'ja' if p['direct_upload'] else 'nee'} | {'; '.join(p['redenen']) or '-'} |")
    w("\n## Aanbevolen aanpak\n")
    if strat["approach"] == "direct":
        w("**Direct uploaden.** Upload de PDF's rechtstreeks en roep de methodiek aan.")
    elif strat["approach"] == "tekstpakket":
        w("**Tekstpakket + steekproef.** Te groot als PDF, maar tekst + outline + "
          "sleutelpagina's passen vermoedelijk in een gesprek. Draai `prepare`.")
    else:
        w("**Veld-voor-veld / via API.** Corpus overschrijdt de context. Review per "
          "veld of laat `review` automatisch chunken (map-reduce).")
    scanned = [p["name"] for p in strat["per_pdf"] if not p["has_text"]]
    if scanned:
        w(f"\n- LET OP geen tekstlaag in: {', '.join(scanned)} -> OCR nodig vooraf.")
    return "\n".join(out)


# --------------------------------------------------------------- scans
def run_scans(corpus, all_pages):
    out = []
    def w(s): out.append(s)
    w("# Geautomatiseerde scans (tekstlaag)\n")
    volts = sorted(set(re.findall(r"\b(\d{2,3})\s?kV\b", corpus)), key=int)
    w(f"## Spanningsniveaus\n- kV-niveaus: {volts}\n")
    w("## Statusvlaggen")
    for term in ["Definitief", "Concept", "Voorlopig", "Ter Acceptatie", "Vervallen"]:
        hits = [(f, pg) for f, pg, p in all_pages if term in p]
        if hits:
            sample = ", ".join(f"{f}:p{pg}" for f, pg in hits[:6])
            w(f"- **{term}**: {len(hits)} pagina's (bv. {sample})")
    w("")
    w("## Open punten / opsteller-notities (uniek, eerste 40)")
    seen = set()
    flag = re.compile(r"nog niet|wordt .* aangepast|nog te |dient nog|verwijderd|"
                      r"toegevoegd|gecontroleerd", re.I)
    for f, pg, p in all_pages:
        for line in p.splitlines():
            l = re.sub(r"\s{2,}", " ", line.strip())
            m = re.match(r"^(0?\d{1,2})\s+(\d+\.\d+)\s+(.{15,})$", l)
            if m and flag.search(m.group(3)):
                key = m.group(3)[:60]
                if key not in seen:
                    seen.add(key)
                    w(f"- {f} p{pg} [{m.group(2)}] {m.group(3)[:130]}")
            if len(seen) >= 40:
                break
    w("")
    docnums = sorted(set(re.findall(r"\b\d{4,6}-\d{2}-\d{4,5}\b", corpus)))
    urls = sorted(set(re.findall(r"https?://\S+", corpus)))
    w("## Verwijzingen (mogelijk niet-aangeleverd)")
    w(f"- {len(docnums)} unieke documentnummers aangehaald.")
    for u in urls[:8]:
        w(f"    - {u[:90]}")
    w("")
    di = set(re.findall(r"\bDI_\d{3}_\d{3}\b", corpus))
    do = set(re.findall(r"\bDO_\d{3}_\d{3}\b", corpus))
    ai = set(re.findall(r"\bAI_\d{3}_\d{3}\b", corpus))
    w("## Signaalinventaris")
    w(f"- DI={len(di)}  DO={len(do)}  AI={len(ai)}  totaal uniek={len(di|do|ai)}\n")
    dates = sorted(set(re.findall(r"\b\d{1,2}-\d{1,2}-20\d{2}\b", corpus)))
    w("## Data-aanduidingen")
    w(f"- {len(dates)} unieke data; bereik {dates[0] if dates else '-'} .. "
      f"{dates[-1] if dates else '-'}")
    return "\n".join(out)


# --------------------------------------------------------------- prepare
def build_candidates(toc, page_bonus, pages_text):
    """Gescoorde sleutelpagina-kandidaten. page_bonus: {pg: bonus}.
    Gebruikt outline-titels (incl. sectie-eerste-pagina's); valt terug op paginatekst
    als er geen outline is. Geeft [(score, pg, label), ...] gesorteerd op score."""
    cand = {}  # pg -> [score, label]

    def bump(pg, score, label):
        if pg in cand:
            if score > cand[pg][0]:
                cand[pg] = [score, label]
        else:
            cand[pg] = [score, label]

    if toc:
        for level, title, pg in toc:
            if not pg or pg < 1:
                continue
            base = _score_title(title)
            if level == 1:
                base += 3  # eerste blad van elk veld/sectie altijd meenemen
            if base > 0:
                bump(pg, base, title)
    else:
        for pg, text in pages_text:
            if _key_match(text, KEY_PATTERNS_STRONG):
                bump(pg, 3, "(uit paginatekst)")
        if pages_text:
            bump(pages_text[0][0], 2, "(eerste pagina)")

    for pg in list(cand):
        cand[pg][0] += page_bonus.get(pg, 0)
    out = [(sc, pg, lbl) for pg, (sc, lbl) in cand.items()]
    out.sort(key=lambda x: (-x[0], x[1]))
    return out


def section_ranges(toc, n, max_pages):
    """Splits op veld-/sectiegrenzen (top-level outline); subsplit als een veld
    groter is dan max_pages. Zonder outline: vaste brokken van max_pages."""
    starts = sorted({pg for lvl, t, pg in (toc or []) if lvl == 1 and 1 <= pg <= n})
    if not starts or starts[0] != 1:
        starts = sorted(set([1] + starts))
    ranges = []
    for i, s in enumerate(starts):
        e = (starts[i + 1] - 1) if i + 1 < len(starts) else n
        for cs in range(s, e + 1, max_pages):
            ranges.append((cs, min(cs + max_pages - 1, e)))
    return ranges


def extra_analyses(all_pages):
    """Titelblok-extractie (heuristisch) + signaal-sluitcontrole, als markdown."""
    rev_re = re.compile(r"\b[Rr]ev(?:isie)?\.?\s*[:=]?\s*([A-Z]?\d{1,2}[A-Z]?)\b")
    blad_re = re.compile(r"\bblad\s*(\d+)(?:\s*(?:/|van)\s*(\d+))?", re.I)
    docnum_re = re.compile(r"\b\d{4,6}-\d{2}-\d{4,5}\b")
    status_words = ["Definitief", "Ter Acceptatie", "Voorlopig", "Concept", "Vervallen"]
    sig_re = re.compile(r"\b[DA][IO]_\d{3}_\d{3}\b")

    out = []
    def w(s): out.append(s)

    # --- titelblok per pagina (alleen waar een tekeningnummer of revisie te vinden is) ---
    w("## Titelblok-extractie (heuristisch)\n")
    w("| Deel | Pagina | Tekeningnr | Revisie | Status | Blad |")
    w("|---|---:|---|---|---|---|")
    rows = 0
    for f, pg, text in all_pages:
        docnum = docnum_re.search(text or "")
        rev = rev_re.search(text or "")
        if not (docnum or rev):
            continue
        status = next((s for s in status_words if s in (text or "")), "")
        blad = blad_re.search(text or "")
        bl = (blad.group(1) + ("/" + blad.group(2) if blad and blad.group(2) else "")) if blad else ""
        w(f"| {f} | {pg} | {docnum.group(0) if docnum else ''} | "
          f"{rev.group(1) if rev else ''} | {status} | {bl} |")
        rows += 1
        if rows >= 200:
            w("| ... | | | | | (afgekapt op 200 regels) |")
            break
    if rows == 0:
        w("| (geen herkenbaar titelblok in de tekstlaag) | | | | | |")

    # --- signaal-sluitcontrole: elk DI_/DO_/AI_-signaal met al zijn vindplaatsen ---
    occ = {}
    for f, pg, text in all_pages:
        for m in set(sig_re.findall(text or "")):
            occ.setdefault(m, []).append(f"{f}:p{pg}")
    w("\n## Signaal-sluitcontrole (DI_/DO_/AI_)\n")
    w(f"- Unieke signalen: **{len(occ)}**")
    single = sorted(s for s, v in occ.items() if len(v) == 1)
    w(f"- Signalen met **slechts één vindplaats** (mogelijk niet gesloten — staat bv. wel in de "
      f"IO-lijst maar niet in een schema, of omgekeerd): **{len(single)}**\n")
    if single:
        w("| Signaal | Enige vindplaats |")
        w("|---|---|")
        for s in single[:100]:
            w(f"| {s} | {occ[s][0]} |")
        if len(single) > 100:
            w(f"| ... | (+{len(single)-100} meer) |")
    w("\n_Let op: dit is een tekstgebaseerde indicatie; een echte sluitcontrole vergt het "
      "onderscheid tussen IO-lijst en schema's. Gebruik dit als startpunt._")
    return "\n".join(out)


def do_prepare(in_dir, out_dir, cfg, emit=print):
    """Bouwt het review_pakket. emit(str) ontvangt voortgangsregels."""
    need_fitz()
    os.makedirs(out_dir, exist_ok=True)
    for sub in ("corpus", "outline", "scans", "split"):
        os.makedirs(os.path.join(out_dir, sub), exist_ok=True)

    inv = inventory(in_dir, cfg)
    strat = decide_strategy(inv, cfg)
    emit(f"Modus: {inv['mode']} | {inv['total_pages']} pagina's | aanpak: {strat['approach']}")
    ocr_enabled = cfg.get("ocr", True)
    ocr_pages = 0

    all_pages, key_global, velden = [], [], []
    pdf_jobs = []
    for idx, it in enumerate(inv["pdfs"], 1):
        emit(f"[{idx}/{len(inv['pdfs'])}] {it['name']} - tekst + outline ...")
        d = fitz.open(it["path"])
        toc = d.get_toc()
        # paginaformaat-bonus + (eventueel) OCR bij lege tekstlaag
        page_bonus, pages_text = {}, []
        txt_path = os.path.join(out_dir, "corpus", f"deel_{idx}.txt")
        with open(txt_path, "w", encoding="utf-8") as fo:
            for pi in range(d.page_count):
                page = d[pi]
                t = page.get_text("text")
                if not t.strip() and ocr_enabled and _ocr_available():
                    try:
                        tp = page.get_textpage_ocr(language="nld+eng", dpi=200, full=True)
                        t = page.get_text("text", textpage=tp)
                        ocr_pages += 1
                    except Exception:
                        pass
                all_pages.append((f"deel_{idx}", pi + 1, t))
                pages_text.append((pi + 1, t))
                page_bonus[pi + 1] = _page_bonus(page)
                fo.write(t)
                if pi < d.page_count - 1:
                    fo.write("\f")
        # outline wegschrijven + velden (top-level) verzamelen
        with open(os.path.join(out_dir, "outline", f"outline_deel_{idx}.txt"),
                  "w", encoding="utf-8") as fo:
            for level, title, pg in toc:
                fo.write(f"{'  '*(level-1)}[{pg}] {title}\n")
                if level == 1:
                    velden.append(f"deel_{idx} - {title}")
        if not any(v.startswith(f"deel_{idx} ") for v in velden):
            velden.append(f"deel_{idx} - {it['name']} (geen outline)")
        # gescoorde kandidaten
        cands = build_candidates(toc, page_bonus, pages_text)
        bron = "outline" if toc else ("paginatekst" if cands else "geen")
        emit(f"     sleutelpagina-kandidaten: {len(cands)} (uit {bron})")
        pdf_jobs.append({"idx": idx, "path": it["path"], "cands": cands})
        # splitsen op veld-/sectiegrenzen indien nodig
        if not strat["per_pdf"][idx - 1]["direct_upload"]:
            emit(f"     splitsen op veldgrenzen (<= {cfg['max_pages_per_pdf']} pagina's) ...")
            base = Path(it["name"]).stem
            for cs, ce in section_ranges(toc, d.page_count, cfg["max_pages_per_pdf"]):
                outd = fitz.open()
                outd.insert_pdf(d, from_page=cs - 1, to_page=ce - 1)
                outd.save(os.path.join(out_dir, "split", f"{base}__p{cs}-{ce}.pdf"))
                outd.close()
        d.close()
    if ocr_pages:
        emit(f"OCR toegepast op {ocr_pages} gescande pagina('s).")
    elif any(not p["has_text"] for p in inv["pdfs"]) and not _ocr_available():
        emit("LET OP: gescande pagina's zonder tekstlaag gevonden, maar Tesseract-OCR is "
             "niet beschikbaar. Installeer Tesseract voor automatische OCR.")

    # --- eerlijke verdeling van het rasterbudget over de PDF's (round-robin, hoogste score eerst) ---
    budget = min(cfg["max_raster_pages"], cfg.get("max_images", cfg["max_raster_pages"]))
    if budget < cfg["max_raster_pages"]:
        emit(f"Sleutelpagina's begrensd op {budget} (max afbeeldingen voor deze AI).")
    queues = [list(j["cands"]) for j in pdf_jobs]
    alloc = {}
    while budget > 0 and any(queues):
        for qi, q in enumerate(queues):
            if not q or budget <= 0:
                continue
            sc, pg, title = q.pop(0)
            alloc.setdefault(pdf_jobs[qi]["idx"], []).append((pg, title))
            budget -= 1
    for job in pdf_jobs:
        items = alloc.get(job["idx"])
        if not items:
            continue
        d = fitz.open(job["path"])
        for pg, title in sorted(items):
            try:
                pix = d[pg - 1].get_pixmap(dpi=cfg["raster_dpi"])
                pix.save(os.path.join(out_dir, "scans", f"deel{job['idx']}_p{pg}.jpg"))
                key_global.append({"deel": job["idx"], "page": pg, "titel": title})
            except Exception as e:
                emit(f"     raster deel{job['idx']} p{pg} mislukt: {e}")

    corpus = "\n".join(p for _, _, p in all_pages)
    scans_md = run_scans(corpus, all_pages) + "\n\n" + extra_analyses(all_pages)
    with open(os.path.join(out_dir, "SCANS.md"), "w", encoding="utf-8") as f:
        f.write(scans_md)
    with open(os.path.join(out_dir, "00_TRIAGE.md"), "w", encoding="utf-8") as f:
        f.write(render_triage(inv, strat, cfg))
    _write_manifest(out_dir, inv, strat, key_global, cfg)
    emit(f"Klaar. Pakket: {out_dir}  ({len(all_pages)} pagina's, "
         f"{len(key_global)} sleutelpagina's, {len(velden)} velden)")
    cfg = dict(cfg)
    cfg["_velden"] = velden
    bundle = do_bundle(out_dir, cfg, emit=emit)
    return {"total_pages": inv["total_pages"], "mode": inv["mode"],
            "approach": strat["approach"], "key_pages": len(key_global),
            "velden": len(velden), "ocr_pages": ocr_pages, "bundle": bundle}


def _write_manifest(out_dir, inv, strat, key_pages, cfg):
    L = []
    def w(s): L.append(s)
    w("# MANIFEST - hoe gebruik je dit review_pakket\n")
    w(f"Aanpak: **{strat['approach']}** - {inv['total_pages']} pagina's - modus {inv['mode']}\n")
    w("## Inhoud")
    w("- `corpus/deel_N.txt` - tekstlaag per PDF")
    w("- `outline/outline_deel_N.txt` - veld-/documenttype-register met paginanummers")
    w(f"- `scans/` - {len(key_pages)} gerasterde sleutelpagina's")
    w("- `split/` - PDF's in upload-bare brokken (indien nodig)")
    w("- `SCANS.md` - geautomatiseerde tekstbevindingen\n")
    w("## Opdracht")
    w("De volledige reviewopdracht (inclusief Excel-output via `Bevindingen_template.xlsx` "
      "en de eis van een review over de volledige breedte van het pakket) staat boven in "
      "`upload/01_LEESMIJ.md`. Upload het hele `upload/`-pakket en plak die opdracht.")
    with open(os.path.join(out_dir, "MANIFEST.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(L))


# --------------------------------------------------------------- bundle (<= N upload-bestanden)
def _read(p):
    return Path(p).read_text(encoding="utf-8", errors="replace") if os.path.isfile(p) else ""


def _pack_blocks(blocks, n_files):
    """Verdeel (label, tekst)-blokken over n_files bestanden, gebalanceerd op grootte."""
    total = sum(len(b) for _, b in blocks)
    target = max(1, math.ceil(total / n_files))
    files, cur, size = [], [], 0
    for label, content in blocks:
        if cur and size + len(content) > target and len(files) < n_files - 1:
            files.append(cur); cur, size = [], 0
        cur.append((label, content)); size += len(content)
    if cur:
        files.append(cur)
    return files


OPDRACHT_KERN = """Je bent een senior ontwerpverificateur secundair voor hoogspanningsstations. Voer een ONAFHANKELIJKE ontwerpverificatie uit over de VOLLEDIGE BREEDTE van dit pakket: alle velden, alle deelbestanden (tekst_*.md), de outline en de sleutelpagina's (sleutelpaginas*.pdf). Beoordeel elk veld en elk documenttype. Werk NIET met een steekproef en sla niets over.

Strikte regels:
- Gebruik uitsluitend de aangeleverde stukken; geen externe aannames of algemene kennis.
- Geef nooit "Voldoet" zonder herleidbaar bewijs; verwijs concreet naar deel/pagina/blad/revisie.
- Benoem ontbrekend bewijs expliciet en claim niet meer dekking dan je daadwerkelijk hebt onderzocht.
- Controleer minimaal: eisconformiteit, interne consistentie, raakvlakken tussen velden, signaal-/IO-sluitcontrole (DI_/DO_/AI_), statusvlaggen (bv. Concept in een definitieve set), uitschakelmatrix en vergrendelingen, en verwijzingen naar niet-aangeleverde documenten.
- Werk de delen systematisch één voor één af en houd een dekkingslog bij (welk deel/veld wel/niet onderzocht).

Lever het resultaat als EXCEL-bestand (.xlsx). Gebruik het meegeleverde sjabloon 'Bevindingen_template.xlsx' en vul exact die tabbladen en kolommen in: 'Samenvatting' (incl. vrijgaveadvies: Vrijgeven / Vrijgeven met voorwaarden / Niet vrijgeven), 'Bevindingen' (ID, Veld, Deel, Pagina/Blad, Documenttype, Categorie, Bevinding, Onderbouwing, Prioriteit Kritiek/Hoog/Middel/Laag, Aanbeveling, Status), 'Open punten', 'Verwijzingen buiten set', 'Signalen', en 'Dekking'. Verzin geen extra kolommen; houd de structuur identiek. Eindig met een expliciete dekkings- en methodeverantwoording (deze review is op de tekstlaag gebaseerd; visuele inspectie alleen voor de sleutelpagina's)."""

OPDRACHT_AFLEVER = {
    "claude": "Aflevering (Claude): maak het Excel-bestand met je bestands-/codefunctionaliteit en bied het als download aan. Lees alle tekst_*.md volledig in voordat je begint. Is het geheel te groot voor één keer, verwerk dan per veld en voeg alles samen tot één werkboek met dezelfde tabbladen.",
    "chatgpt": "Aflevering (ChatGPT): gebruik je data-analyse (Python/openpyxl) om een ECHT .xlsx-bestand te genereren en als download aan te bieden. Verwerk elk geüpload tekstbestand volledig — vat niet samen op basis van een fragment. Bij omvang: itereer per deel en bouw het werkboek incrementeel verder; behoud alle eerdere bevindingen.",
    "copilot": "Aflevering (Copilot): lever het resultaat als Excel-werkboek (of als tabellen die je via 'Exporteren naar Excel' opslaat in de structuur van het sjabloon). Copilot kan minder bestanden per gesprek aan: upload zo nodig per veld en laat Copilot de bevindingen CUMULATIEF in hetzelfde werkboek bijhouden, zodat de volledige breedte alsnog wordt afgedekt.",
}


def build_opdracht(ai_key, ai_label):
    aflever = OPDRACHT_AFLEVER.get(ai_key, OPDRACHT_AFLEVER["claude"])
    return f"# Opdracht voor {ai_label}\n\n{OPDRACHT_KERN}\n\n{aflever}"


def do_bundle(pakket, cfg, emit=print, max_files=None):
    """Vouwt review_pakket samen tot <= max_files upload-bare bestanden in pakket/upload/."""
    need_fitz()
    max_files = int(max_files or cfg.get("max_upload_files", 20))
    if max_files < 3:
        max_files = 3
    up = os.path.join(pakket, "upload")
    shutil.rmtree(up, ignore_errors=True)
    os.makedirs(up)

    # --- vaste bestanden: leesmij (opdracht + triage + manifest + methodiek) ---
    ai = cfg.get("ai_label") or "je AI-chat"
    ai_key = cfg.get("ai_key") or "claude"
    opdracht = build_opdracht(ai_key, ai)
    leesmij = (opdracht + "\n\n---\n\n# Over dit pakket\n\nUpload alle bestanden in deze map (`upload/`) "
               "in EEN " + ai + "-gesprek. De tekst van alle delen is samengevoegd in de "
               "`tekst_*.md`-bestanden (met outline per deel); eventuele sleutelpagina's staan in "
               "`sleutelpaginas*.pdf`; vul het Excel-sjabloon `Bevindingen_template.xlsx` in.\n\n"
               "---\n\n" + _read(os.path.join(pakket, "MANIFEST.md")) +
               "\n\n---\n\n" + _read(os.path.join(pakket, "00_TRIAGE.md")))
    methodiek = load_methodiek()
    if methodiek:
        leesmij += "\n\n---\n\n# Methodiek voor de beoordeling\n" + methodiek
    Path(os.path.join(up, "01_LEESMIJ.md")).write_text(leesmij, encoding="utf-8")
    overhead = 1
    if os.path.isfile(os.path.join(pakket, "SCANS.md")):
        shutil.copyfile(os.path.join(pakket, "SCANS.md"), os.path.join(up, "02_SCANS.md"))
        overhead += 1
    # Excel-sjabloon meeleveren; Dekking-tabblad voorinvullen met alle velden/delen
    tmpl = os.path.join(app_dir(), "templates", "Bevindingen_template.xlsx")
    dest_xlsx = os.path.join(up, "Bevindingen_template.xlsx")
    velden = cfg.get("_velden") or []
    if os.path.isfile(tmpl):
        filled = False
        if velden:
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, Border, Side
                wb = load_workbook(tmpl)
                ws = wb["Dekking"]
                thin = Side(style="thin", color="BFBFBF")
                bd = Border(left=thin, right=thin, top=thin, bottom=thin)
                r = 3  # rij 1 = notitie, rij 2 = kop
                for v in velden:
                    ws.cell(r, 1, v).font = Font(name="Arial", size=10)
                    ws.cell(r, 2, "").font = Font(name="Arial", size=10)  # onderzocht ja/nee
                    ws.cell(r, 3, "tekstlaag")
                    for c in range(1, 6):
                        ws.cell(r, c).border = bd
                    r += 1
                wb.save(dest_xlsx)
                filled = True
            except Exception as e:
                emit(f"     Dekking voorinvullen overgeslagen ({e}); sjabloon gekopieerd.")
        if not filled:
            shutil.copyfile(tmpl, dest_xlsx)
        overhead += 1

    # --- sleutelpagina's bundelen tot PDF('s) van <= max_pages_per_pdf pagina's ---
    imgs = sorted(glob.glob(os.path.join(pakket, "scans", "*.jpg")) +
                  glob.glob(os.path.join(pakket, "scans", "*.png")))
    per_pdf = cfg["max_pages_per_pdf"]
    scan_pdfs = 0
    for gi, start in enumerate(range(0, len(imgs), per_pdf), 1):
        group = imgs[start:start + per_pdf]
        doc = fitz.open()
        for ip in group:
            im = fitz.open(ip)
            doc.insert_pdf(fitz.open("pdf", im.convert_to_pdf()))
            im.close()
        name = "03_sleutelpaginas.pdf" if len(imgs) <= per_pdf else f"03_sleutelpaginas_{gi}.pdf"
        doc.save(os.path.join(up, name)); doc.close()
        scan_pdfs += 1
    overhead += scan_pdfs

    # --- tekst + outline samenvoegen en over de resterende slots verdelen ---
    text_slots = max(1, max_files - overhead)
    blocks = []
    for txt in sorted(glob.glob(os.path.join(pakket, "corpus", "deel_*.txt"))):
        deel = os.path.splitext(os.path.basename(txt))[0]
        outline = _read(os.path.join(pakket, "outline", f"outline_{deel}.txt"))
        blocks.append((deel + ":outline",
                       f"\n\n########## {deel} ##########\n### OUTLINE\n{outline}\n### TEKST\n"))
        for pi, page in enumerate(load_pages(txt), 1):
            blocks.append((f"{deel}:p{pi}", f"\n--- {deel} pagina {pi} ---\n{page}"))

    total_chars = sum(len(b) for _, b in blocks)
    budget = cfg.get("upload_chars_per_file", 500_000)
    n_by_budget = max(1, math.ceil(total_chars / budget))
    n_files = min(text_slots, n_by_budget)
    warn = n_by_budget > text_slots

    packed = _pack_blocks(blocks, n_files)
    width = max(2, len(str(len(packed))))
    for i, grp in enumerate(packed, 1):
        body = "".join(c for _, c in grp)
        Path(os.path.join(up, f"tekst_{str(i).zfill(width)}.md")).write_text(body, encoding="utf-8")

    n_total = len([f for f in os.listdir(up)])
    est = est_tokens(total_chars)
    n_images = len(imgs)
    max_img = cfg.get("max_images", 20)
    emit(f"Upload-bundel: {n_total} bestand(en) in {up} (limiet {max_files}).")
    emit(f"  tekst {n_files} bestand(en), scans {scan_pdfs} PDF met {n_images} afbeelding(en) "
         f"(limiet {max_img}), ~{est:,} tokens tekst.")
    if n_images > max_img:
        emit(f"  LET OP: {n_images} afbeeldingen > limiet {max_img}; verlaag 'max afbeeldingen'.")
    if warn:
        emit(f"  LET OP: de tekst (~{est:,} tokens) is groter dan {n_files} x {budget//1000}k "
             "tekens; deze past waarschijnlijk niet in EEN gesprek (contextvenster). "
             "Overweeg de delen veld-voor-veld in losse gesprekken te uploaden.")
    return {"upload_dir": up, "n_files": n_total, "text_files": n_files,
            "scan_pdfs": scan_pdfs, "n_images": n_images, "max_images": max_img,
            "est_tokens": est, "fits_context_warning": warn,
            "max_files": max_files, "opdracht": opdracht}


# --------------------------------------------------------------- methodiek
def load_methodiek(base=None):
    md_dir = os.path.join(base or app_dir(), "methodiek")
    if not os.path.isdir(md_dir):
        return ""
    parts = []
    for name in sorted(os.listdir(md_dir)):
        if name.endswith(".md"):
            with open(os.path.join(md_dir, name), encoding="utf-8") as f:
                parts.append(f"\n\n===== {name} =====\n\n" + f.read())
    return "".join(parts)

# --------------------------------------------------------------- CLI
def make_cfg(args):
    cfg = dict(DEFAULTS)
    for k in ("max_file_mb", "max_pages_per_pdf", "raster_dpi", "max_raster_pages",
              "max_images", "max_upload_files"):
        v = getattr(args, k, None)
        if v is not None:
            cfg[k] = v
    return cfg


def cmd_triage(args):
    cfg = make_cfg(args)
    inv = inventory(args.in_dir, cfg)
    rep = render_triage(inv, decide_strategy(inv, cfg), cfg)
    print("\n" + rep + "\n")
    if getattr(args, "out", None):
        os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
        Path(args.out).write_text(rep, encoding="utf-8")


def cmd_prepare(args):
    do_prepare(args.in_dir, args.out, make_cfg(args))


def cmd_bundle(args):
    do_bundle(args.pakket, dict(DEFAULTS), max_files=args.max_upload_files)


def build_parser():
    ap = argparse.ArgumentParser(description="Voorbereiding van secundaire ontwerppakketten voor review in Claude/ChatGPT/Copilot.")
    sub = ap.add_subparsers(dest="cmd")

    def common(p):
        p.add_argument("--in", dest="in_dir", default="./pdfs")
        p.add_argument("--max-file-mb", type=float, dest="max_file_mb")
        p.add_argument("--max-pages-per-pdf", type=int, dest="max_pages_per_pdf")
        p.add_argument("--raster-dpi", type=int, dest="raster_dpi")
        p.add_argument("--max-raster-pages", type=int, dest="max_raster_pages")
        p.add_argument("--max-images", type=int, dest="max_images")
        p.add_argument("--max-upload-files", type=int, dest="max_upload_files")

    p1 = sub.add_parser("triage"); common(p1)
    p1.add_argument("--out"); p1.set_defaults(func=cmd_triage)
    p2 = sub.add_parser("prepare"); common(p2)
    p2.add_argument("--out", default="./review_pakket"); p2.set_defaults(func=cmd_prepare)
    pb = sub.add_parser("bundle")
    pb.add_argument("--pakket", default="./review_pakket")
    pb.add_argument("--max-upload-files", type=int, dest="max_upload_files", default=20)
    pb.set_defaults(func=cmd_bundle)
    return ap


def main():
    if len(sys.argv) == 1:
        print(textwrap.dedent("""
            Ontwerpchecker - gebruik:
              python ontwerpchecker.py triage  --in ./pdfs
              python ontwerpchecker.py prepare --in ./pdfs --out ./review_pakket
              python ontwerpchecker.py bundle  --pakket ./review_pakket --max-upload-files 20
            Of start het bedieningspaneel:  python webapp.py
        """))
        return
    args = build_parser().parse_args()
    if not getattr(args, "cmd", None):
        build_parser().print_help(); return
    args.func(args)


if __name__ == "__main__":
    main()
